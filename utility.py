from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

def init() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trainees"
    sheet["A1"], sheet["B1"], sheet["C1"] = "id", "CourseId", "Name"
    sheet["D1"], sheet["E1"] = "Background/degree", "WorkExperience"
    workbook.create_sheet("CourseDetails")
    sheet = workbook["CourseDetails"]
    sheet["A1"], sheet["B1"], sheet["C1"] = "CourseId", "trainerId", "Description"
    workbook.create_sheet("Trainers")
    sheet = workbook["Trainers"]
    sheet["A1"], sheet["B1"] = "id", "managerid"
    sheet["C1"], sheet["D1"], sheet["E1"] = "name", "email", "phone number"
    workbook.create_sheet("Managers")
    sheet = workbook["Managers"]
    sheet["A1"], sheet["B1"] = "id", "name"
    sheet["C1"], sheet["D1"] = "email", "phone number"
    workbook.save(filename="course_manager.xlsx")
def selectColumns(sheet: Worksheet) -> list:
    for cell in sheet[1]:
        print(cell.column)
    return [cell.column for cell in sheet[1] if input(f"Display {cell.value}? y / n ").casefold() == "y"]
def join(sheetA: Worksheet, sheetB: Worksheet, Acolumns: list, Bcolumns: list, workbook: Workbook, title: str):
    try:
        workbook.remove(workbook[title])
    except KeyError:
        pass
    workbook.create_sheet(title)
    sheetAB = workbook[title]
    Arow = [sheetA.cell(column=col, row = 1).value for col in Acolumns]
    Brow = [sheetB.cell(column=col, row = 1).value for col in Bcolumns]
    sheetAB.append([*Arow, *Brow])
    for cell in sheetA["B"]:
        if cell.value in [cell2.value for cell2 in sheetB["A"][1:]]:
            Arow = [sheetA.cell(column=col, row = cell.row).value for col in Acolumns]
            Brow = [sheetB.cell(column=col, row = cell.row).value for col in Bcolumns]
            sheetAB.append([*Arow, *Brow])
def setAttendance(sheet: Worksheet, absent: list):
    notExist = [x for x in absent if x not in [cell.value for cell in sheet["A"]]]
    if notExist:
        print("These ids were not found", " ".join(notExist))
    for row in sheet.rows:
        if row[0].value in absent:
            row[2].value = "A"
def sendEmail(workbook: Workbook, title: str):
    try:
       workbook[title]
    except KeyError:
        return
    courseId = title.split(" ")[1]
    trainerId = None
    managerId = None
    for cell in workbook["CourseDetails"]["A"]:
        if courseId == cell.value:
            trainerId = workbook["CourseDetails"][f"B{cell.row}"]
    for cell in workbook["Trainers"]["A"]:
        if trainerId == cell.value:
            managerId = workbook["Trainers"][f"B{cell.row}"]
    for cell in workbook["Managers"]["A"]:
        if managerId == cell.value:
            return workbook["Managers"][f"B{cell.row}"]
    

    