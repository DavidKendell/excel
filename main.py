from openpyxl import Workbook
from Controller import *
from openpyxl import load_workbook
from utility import *
import os



if not os.path.isfile("course_manager.xlsx"):
    init()
workbook = load_workbook(filename="course_manager.xlsx")
sheet = workbook["Trainees"]
choice = None
date = None
controller = Controller(sheet)
while choice != "q":
    choice = input("""
    1 Show sheets
    2 Set active sheet
    3 Create entry
    4 Read entry
    5 Update entry
    6 Delete entry
    7 Add session
    q Quit
    """)
    match(choice):
        case "1":
            print(", ".join(workbook.sheetnames))
        case "2":
            try:
                sheetname = input("Enter name of sheet ")
                if sheetname.find("Map") != -1:
                    print("Map sheets are auto-generated")
                    continue
                sheet = workbook[sheetname]
                controller.data = sheet
            except KeyError:
                print("No such sheet")
                continue
        case "3":
            newId = input("Enter the id ")
            if controller.find(newId) != 0:
                print("id must be unique")
                continue
            controller.add(newId, [input(f"Enter the value for {cell.value} ") for cell in sheet[1][1:]])
        case "4":
            row = controller.find(input("Enter the id "))
            if not row:
                print("Not found ")
                continue
            sizes = [len(cell.value) for cell in sheet[1] + sheet[row]]
            width = max(sizes)
            print("|".join([f"{cell.value:^{width}}" for cell in sheet[1]]))
            print("-"*(len(sheet[1])*(width+1)-1))
            print("|".join([f"{cell.value:^{width}}" for cell in sheet[row]]))
        case "5":
            row = controller.find(input("Enter the id "))
            if not row:
                print("id not found")
                continue
            sizes = [len(cell.value) for cell in sheet[1] + sheet[row]]
            width = max(sizes)
            print("|".join([f"{cell.value:^{width}}" for cell in sheet[1]]))
            print("-"*(len(sheet[1])*(width+1)-1))
            print("|".join([f"{cell.value:^{width}}" for cell in sheet[row]]))
            controller.update(row, [input(f"Enter the new value for {cell.value}. Enter blank string to skip ") for cell in sheet[1][1:]])
            print("|".join([f"{cell.value:^{width}}" for cell in sheet[1]]))
            print("-"*(len(sheet[1])*(width+1)-1))
            print("|".join([f"{cell.value:^{width}}" for cell in sheet[row]]))
        case "6":
            controller.delete(input("Enter the id "))
        case "7":
            date = input("Enter date ")
            courseId = input("Enter course id")
            workbook.create_sheet(date + " " + courseId)
            attendance = workbook[date + " " + courseId]
            attendance["A1"], attendance["B1"], attendance["C1"] = "id", "name", "present/absent"
            trainees = workbook["Trainees"]
            rows = trainees.rows
            next(rows)
            for row in rows:
                if row[1].value == courseId:
                    attendance.append([row[0].value, row[2].value, "P"])
            setAttendance(attendance, [input("Enter id of absent trainee ") for i in range(int(input("Enter number of absent students ")))])

            
        case "q":
            break
email = sendEmail(workbook ,date + " " + "c343")
print(f"{email} has been notified about c343 session dated {date}")

print("Joining Trainees with CourseDetails")
print("Select Trainees columns to display")
A = selectColumns(workbook["Trainees"])
print("Select CourseDetails columns to display")
B = selectColumns(workbook["CourseDetails"])
join(workbook["Trainees"], workbook["CourseDetails"], A, B, workbook, "TraineeCourseMap")
print("Joining Trainers with CourseDetails")
print("Select Trainers columns to display")
A = selectColumns(workbook["Trainers"])
print("Select CourseDetails columns to display")
B = selectColumns(workbook["CourseDetails"])
join(workbook["CourseDetails"], workbook["Trainers"], B, A, workbook, "TrainerCourseMap")
workbook.save(filename="course_manager.xlsx")